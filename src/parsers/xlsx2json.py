import os
import typing

import openpyxl
from tqdm import tqdm

from src.base.parsers import BaseParser


class CellParser:
    """
    Parse cell information from sheet
    """
    def __init__(self,
                 row: int,
                 col: int,
                 description: str,
                 alias: str,
                 cell_type: str = "str"):
        self.row = row
        self.col = col
        self.description = description
        self.alias = alias
        self.cell_type = cell_type

    def parse(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict:
        cell = {
            f"{self.alias}": sheet.cell(row=self.row, column=self.col).value
        }

        return cell


class SheetParser:
    """
    Parse all information from sheet
    """
    def __init__(self, cell_parsers: typing.Tuple[CellParser]):
        self.parsers = cell_parsers

    def parse(self, sheet):
        result = dict()
        for parser in self.parsers:
            cell_info = parser.parse(sheet)
            result.update(cell_info)
        return result


class XLReader:
    def __init__(self, path: typing.Union[str, os.PathLike]):
        self.xl: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(path)
        self.path = path

    def get_sheet_iterator(self) -> openpyxl.worksheet.worksheet.Worksheet:
        for sheet in self.xl.sheetnames:
            yield self.xl[sheet]


# List of cell parser used to parse one sheet
DEFAULT_CELL_PARSERS = (
    CellParser(1, 8, "Some strange parameter", "sequence", "int"),
    CellParser(3, 1, "Name prefix", "name_prefix"),
    CellParser(3, 2, "Name Surname", "name_surname"),
    CellParser(3, 8, "Unknown parameter", "unk1"),
    CellParser(5, 1, "Flight", "flight"),
    CellParser(5, 4, "Departure", "departure"),
    CellParser(5, 8, "Arrival", "arrival"),
    CellParser(7, 2, "Gate", "gate"),
    CellParser(7, 4, "Short name of departure airport", "dep_short"),
    CellParser(7, 8, "Short name of arrival airport", "arr_short"),
    CellParser(9, 1, "Departure date", "dep_date", "date"),
    CellParser(9, 3, "Departure time", "dep_time", "time"),
    CellParser(9, 5, "Possibly comment on airlines", "airlines_comment"),
    CellParser(11, 1, "Additional information on flight", "appendix_info"),
    CellParser(11, 8, "Seat", "seat"),
    CellParser(13, 2, "PNR", "pnr"),
    CellParser(13, 4, "Ticket type", "ticket_type"),
    CellParser(13, 5, "Ticket number", "ticket_num"),
)


class XLSXParser(BaseParser):
    """
    Parser from XLSX format presented in that project
    """
    input_format = "xlsx"

    def __init__(self,
                 output_dir: typing.Union[str, os.PathLike],
                 cell_parsers: typing.Tuple[CellParser] = DEFAULT_CELL_PARSERS):
        super().__init__(output_dir)
        self.cell_parsers = cell_parsers
        self.sheetParser = SheetParser(cell_parsers)

    def parse(self, path: typing.Union[str, os.PathLike], *args, **kwargs):
        xl = XLReader(path)

        results = []
        for sheet in xl.get_sheet_iterator():
            sheet_info = self.sheetParser.parse(sheet)
            results.append(sheet_info)

        output_path = kwargs.get('output_path', self._get_default_output_path(path))
        self.to_json(results, output_path)


def parse_to_json(parser: BaseParser, paths: typing.Iterable[str]):
    for path in tqdm(paths):
        assert path.endswith(".xlsx"), f"File must be xlsx, got {path}"
        parser.parse(path)


if __name__ == "__main__":
    import glob

    parser = XLSXParser(output_dir="../../data/xlsx-parsed")
    parse_to_json(parser, glob.glob("data/*.xlsx"))
